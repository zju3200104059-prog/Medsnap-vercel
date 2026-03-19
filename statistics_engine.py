# -*- coding: utf-8 -*-
"""
MedSnap 统计分析引擎 — 纯 Python 计算模块
不依赖 Flask，输入 pandas DataFrame，输出标准化 dict。
"""

import json
import uuid
import io
import copy
import threading
from collections import OrderedDict

import numpy as np
import pandas as pd
from scipy import stats as sp_stats

# ==============================================================================
#  模块级数据集存储
# ==============================================================================

_datasets = {}           # dataset_id → DatasetManager
_datasets_lock = threading.Lock()
MAX_DATASETS = 10
MAX_ROWS = 50000


def get_dataset(dataset_id):
    with _datasets_lock:
        return _datasets.get(dataset_id)


def register_dataset(dm):
    with _datasets_lock:
        if len(_datasets) >= MAX_DATASETS:
            oldest_key = next(iter(_datasets))
            del _datasets[oldest_key]
        _datasets[dm.dataset_id] = dm
    return dm.dataset_id


def remove_dataset(dataset_id):
    with _datasets_lock:
        _datasets.pop(dataset_id, None)


# ==============================================================================
#  DatasetManager — 数据集生命周期管理
# ==============================================================================

class DatasetManager:
    def __init__(self, df, name='unnamed'):
        if len(df) > MAX_ROWS:
            df = df.head(MAX_ROWS)
        self.dataset_id = f"ds_{uuid.uuid4().hex[:8]}"
        self.name = name
        self.df = df
        self._labels = {}        # col → 标签文字
        self._history = []       # DataFrame 栈，用于撤销
        self._max_history = 20

    # ---------- 加载方法 ----------

    @classmethod
    def from_csv(cls, file_obj, filename='upload.csv'):
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'latin-1']
        for enc in encodings:
            try:
                file_obj.seek(0)
                df = pd.read_csv(file_obj, encoding=enc)
                return cls(df, name=filename)
            except (UnicodeDecodeError, pd.errors.ParserError):
                continue
        raise ValueError("无法解析 CSV 文件，请检查编码格式")

    @classmethod
    def from_excel(cls, file_obj, filename='upload.xlsx'):
        file_obj.seek(0)
        df = pd.read_excel(file_obj, engine='openpyxl')
        return cls(df, name=filename)

    @classmethod
    def from_records(cls, rows):
        """从 medical_records 的 extracted_data JSON 列表构建 DataFrame"""
        all_flat = []
        for row in rows:
            raw = row.get('extracted_data', '{}') or '{}'
            try:
                data = json.loads(raw) if isinstance(raw, str) else raw
            except json.JSONDecodeError:
                continue
            flat = _flatten_dict(data)
            all_flat.append(flat)
        if not all_flat:
            raise ValueError("无有效数据记录")
        df = pd.DataFrame(all_flat)
        # 尝试将可转数值的列转为 float
        for col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='ignore')
        return cls(df, name='导入病历数据')

    # ---------- 信息查询 ----------

    def get_info(self):
        df = self.df
        cols_info = []
        for col in df.columns:
            dtype = 'numeric' if pd.api.types.is_numeric_dtype(df[col]) else 'categorical'
            missing = int(df[col].isna().sum())
            missing_pct = round(missing / len(df) * 100, 1) if len(df) > 0 else 0
            unique = int(df[col].nunique())
            cols_info.append({
                'name': col,
                'label': self._labels.get(col, ''),
                'dtype': dtype,
                'missing': missing,
                'missing_pct': missing_pct,
                'unique': unique,
            })
        return {
            'dataset_id': self.dataset_id,
            'name': self.name,
            'rows': len(df),
            'cols': len(df.columns),
            'columns': cols_info,
            'can_undo': len(self._history) > 0,
        }

    def get_preview(self, page=1, page_size=50):
        start = (page - 1) * page_size
        end = start + page_size
        subset = self.df.iloc[start:end]
        records = json.loads(subset.to_json(orient='records', force_ascii=False, date_format='iso'))
        return {
            'columns': list(self.df.columns),
            'data': records,
            'total_rows': len(self.df),
            'page': page,
            'page_size': page_size,
            'total_pages': max(1, (len(self.df) + page_size - 1) // page_size),
        }

    def get_columns_by_type(self):
        df = self.df
        numeric = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        categorical = [c for c in df.columns if not pd.api.types.is_numeric_dtype(df[c])]
        return {'numeric': numeric, 'categorical': categorical, 'all': list(df.columns)}

    # ---------- 历史管理 ----------

    def _push_history(self):
        if len(self._history) >= self._max_history:
            self._history.pop(0)
        self._history.append(self.df.copy())

    def undo(self):
        if not self._history:
            return False
        self.df = self._history.pop()
        return True


# ==============================================================================
#  Preprocessor — 模块 1: 数据预处理
# ==============================================================================

class Preprocessor:

    @staticmethod
    def encode_variable(dm, col, mapping=None):
        """分类变量编码。mapping=None 时自动编码。"""
        dm._push_history()
        df = dm.df
        if mapping:
            df[col] = df[col].map(mapping)
        else:
            codes, uniques = pd.factorize(df[col])
            df[col] = codes
            mapping = {str(u): int(c) for c, u in enumerate(uniques)}
        dm.df = df
        return {'msg': f'已对 {col} 完成编码', 'mapping': mapping}

    @staticmethod
    def set_labels(dm, label_dict):
        dm._labels.update(label_dict)
        return {'msg': f'已更新 {len(label_dict)} 个变量标签'}

    @staticmethod
    def detect_outliers(dm, cols, method='iqr'):
        results = []
        df = dm.df
        for col in cols:
            if not pd.api.types.is_numeric_dtype(df[col]):
                results.append({'variable': col, 'outlier_count': 0, 'msg': '非数值变量，跳过'})
                continue
            series = df[col].dropna()
            if method == 'iqr':
                q1 = series.quantile(0.25)
                q3 = series.quantile(0.75)
                iqr = q3 - q1
                lower = q1 - 1.5 * iqr
                upper = q3 + 1.5 * iqr
            else:  # 3sigma
                mean = series.mean()
                std = series.std()
                lower = mean - 3 * std
                upper = mean + 3 * std
            outlier_mask = (series < lower) | (series > upper)
            count = int(outlier_mask.sum())
            results.append({
                'variable': col,
                'outlier_count': count,
                'lower_bound': round(float(lower), 4),
                'upper_bound': round(float(upper), 4),
                'method': 'IQR箱线图法' if method == 'iqr' else '3σ原则',
            })
        return results

    @staticmethod
    def handle_outliers(dm, cols, method='iqr', action='delete'):
        dm._push_history()
        df = dm.df.copy()
        total_removed = 0
        for col in cols:
            if not pd.api.types.is_numeric_dtype(df[col]):
                continue
            series = df[col]
            if method == 'iqr':
                q1 = series.quantile(0.25)
                q3 = series.quantile(0.75)
                iqr = q3 - q1
                lower, upper = q1 - 1.5 * iqr, q3 + 1.5 * iqr
            else:
                mean, std = series.mean(), series.std()
                lower, upper = mean - 3 * std, mean + 3 * std
            mask = (series < lower) | (series > upper)
            if action == 'delete':
                df = df[~mask]
                total_removed += int(mask.sum())
            elif action == 'clip':
                df[col] = series.clip(lower, upper)
        dm.df = df.reset_index(drop=True)
        return {'msg': f'异常值处理完成，删除/裁剪 {total_removed} 条记录'}

    @staticmethod
    def handle_missing(dm, cols, method='delete'):
        dm._push_history()
        df = dm.df
        before = len(df)
        if method == 'delete':
            df = df.dropna(subset=cols)
        elif method == 'mean':
            for col in cols:
                if pd.api.types.is_numeric_dtype(df[col]):
                    df[col] = df[col].fillna(df[col].mean())
        elif method == 'median':
            for col in cols:
                if pd.api.types.is_numeric_dtype(df[col]):
                    df[col] = df[col].fillna(df[col].median())
        elif method == 'mode':
            for col in cols:
                mode_val = df[col].mode()
                if len(mode_val) > 0:
                    df[col] = df[col].fillna(mode_val.iloc[0])
        dm.df = df.reset_index(drop=True)
        after = len(dm.df)
        return {'msg': f'缺失值处理完成（处理前 {before} 行 → 处理后 {after} 行）'}

    @staticmethod
    def filter_samples(dm, col, condition, value):
        """条件过滤样本。condition: eq/ne/gt/lt/gte/lte/isna/notna"""
        dm._push_history()
        df = dm.df
        before = len(df)
        if condition == 'isna':
            df = df[df[col].notna()]
        elif condition == 'notna':
            df = df[df[col].isna()]
        else:
            try:
                val = float(value) if pd.api.types.is_numeric_dtype(df[col]) else value
            except (ValueError, TypeError):
                val = value
            ops = {'eq': lambda s: s == val, 'ne': lambda s: s != val,
                   'gt': lambda s: s > val, 'lt': lambda s: s < val,
                   'gte': lambda s: s >= val, 'lte': lambda s: s <= val}
            if condition in ops:
                df = df[ops[condition](df[col])]
        dm.df = df.reset_index(drop=True)
        return {'msg': f'样本过滤完成（{before} → {len(dm.df)} 行）'}

    @staticmethod
    def generate_variable(dm, source_cols, operation, new_name):
        """基于现有变量算术运算生成新变量。operation: sum/mean/ratio/diff"""
        dm._push_history()
        df = dm.df
        numeric_df = df[source_cols].apply(pd.to_numeric, errors='coerce')
        if operation == 'sum':
            df[new_name] = numeric_df.sum(axis=1)
        elif operation == 'mean':
            df[new_name] = numeric_df.mean(axis=1)
        elif operation == 'ratio':
            if len(source_cols) == 2:
                denom = numeric_df.iloc[:, 1].replace(0, np.nan)
                df[new_name] = numeric_df.iloc[:, 0] / denom
            else:
                return {'msg': '比值运算需要选择恰好 2 个变量', 'error': True}
        elif operation == 'diff':
            if len(source_cols) == 2:
                df[new_name] = numeric_df.iloc[:, 0] - numeric_df.iloc[:, 1]
            else:
                return {'msg': '差值运算需要选择恰好 2 个变量', 'error': True}
        dm.df = df
        return {'msg': f'已生成新变量 "{new_name}"'}

    @staticmethod
    def standardize(dm, cols, method='zscore'):
        dm._push_history()
        df = dm.df
        for col in cols:
            if not pd.api.types.is_numeric_dtype(df[col]):
                continue
            series = df[col].astype(float)
            if method == 'zscore':
                mean, std = series.mean(), series.std()
                df[col] = (series - mean) / std if std != 0 else 0
            elif method == 'center':
                df[col] = series - series.mean()
            elif method == 'minmax':
                mn, mx = series.min(), series.max()
                df[col] = (series - mn) / (mx - mn) if mx != mn else 0
        dm.df = df
        return {'msg': f'已对 {len(cols)} 个变量完成 {method} 标准化'}


# ==============================================================================
#  DescriptiveAnalyzer — 模块 2: 通用描述性分析
# ==============================================================================

class DescriptiveAnalyzer:

    @staticmethod
    def frequency(df, col):
        """频数分析：频次、累计频次、百分比、累计百分比"""
        vc = df[col].value_counts(dropna=False).sort_index()
        total = vc.sum()
        cum = vc.cumsum()
        rows = []
        for val, cnt in vc.items():
            label = str(val) if pd.notna(val) else '缺失'
            pct = round(cnt / total * 100, 2) if total > 0 else 0
            cum_cnt = int(cum[val])
            cum_pct = round(cum_cnt / total * 100, 2) if total > 0 else 0
            rows.append({
                '值': label, '频次': int(cnt), '累计频次': cum_cnt,
                '百分比(%)': pct, '累计百分比(%)': cum_pct
            })
        # 附带柱状图
        chart = {
            'title': {'text': f'{col} 频数分布', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'category', 'data': [r['值'] for r in rows], 'axisLabel': {'rotate': 30}},
            'yAxis': {'type': 'value', 'name': '频次'},
            'series': [{'type': 'bar', 'data': [r['频次'] for r in rows],
                        'itemStyle': {'color': '#5b8ff9'}, 'label': {'show': True, 'position': 'top'}}],
            'grid': {'bottom': 80}
        }
        return {
            'method_name': '频数分析',
            'summary_table': rows,
            'columns': ['值', '频次', '累计频次', '百分比(%)', '累计百分比(%)'],
            'footnotes': [f'有效样本量 N = {total}'],
            'charts': [chart],
            'interpretation': f'变量 "{col}" 共有 {len(vc)} 个不同取值，总频次 {total}。'
        }

    @staticmethod
    def descriptive_stats(df, cols, group_var=None):
        """描述统计：N、均值、标准差、中位数、最小值、最大值、Q1、Q3"""
        def _calc(series, name=''):
            s = pd.to_numeric(series, errors='coerce').dropna()
            if len(s) == 0:
                return {'变量': name, 'N': 0, '均值': '-', '标准差': '-', '中位数': '-',
                        '最小值': '-', '最大值': '-', 'Q1': '-', 'Q3': '-'}
            return {
                '变量': name, 'N': int(len(s)),
                '均值': round(float(s.mean()), 4),
                '标准差': round(float(s.std()), 4),
                '中位数': round(float(s.median()), 4),
                '最小值': round(float(s.min()), 4),
                '最大值': round(float(s.max()), 4),
                'Q1': round(float(s.quantile(0.25)), 4),
                'Q3': round(float(s.quantile(0.75)), 4),
            }

        rows = []
        box_data = []
        box_labels = []

        if group_var and group_var in df.columns:
            groups = df.groupby(group_var)
            for gname, gdf in groups:
                for col in cols:
                    row = _calc(gdf[col], f'{col} ({gname})')
                    rows.append(row)
                    s = pd.to_numeric(gdf[col], errors='coerce').dropna()
                    if len(s) > 0:
                        box_data.append([float(s.min()), float(s.quantile(0.25)),
                                         float(s.median()), float(s.quantile(0.75)), float(s.max())])
                        box_labels.append(f'{col}({gname})')
        else:
            for col in cols:
                row = _calc(df[col], col)
                rows.append(row)
                s = pd.to_numeric(df[col], errors='coerce').dropna()
                if len(s) > 0:
                    box_data.append([float(s.min()), float(s.quantile(0.25)),
                                     float(s.median()), float(s.quantile(0.75)), float(s.max())])
                    box_labels.append(col)

        chart = {
            'title': {'text': '数据分布（箱线图）', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': box_labels, 'axisLabel': {'rotate': 30}},
            'yAxis': {'type': 'value'},
            'series': [{'type': 'boxplot', 'data': box_data,
                        'itemStyle': {'color': '#dbeafe', 'borderColor': '#2563eb'}}],
            'grid': {'bottom': 80}
        }
        columns = ['变量', 'N', '均值', '标准差', '中位数', '最小值', '最大值', 'Q1', 'Q3']
        return {
            'method_name': '描述统计',
            'summary_table': rows,
            'columns': columns,
            'footnotes': [],
            'charts': [chart],
            'interpretation': f'共分析 {len(cols)} 个变量，{len(rows)} 组结果。'
        }

    @staticmethod
    def cross_tabulation(df, row_var, col_var):
        """交叉列联表 + 卡方检验"""
        ct = pd.crosstab(df[row_var], df[col_var], margins=True, margins_name='合计')
        chi2, p, dof, expected = sp_stats.chi2_contingency(
            pd.crosstab(df[row_var], df[col_var])
        )
        # 构建频数表
        freq_rows = []
        for idx in ct.index:
            row = {'': str(idx)}
            for col_name in ct.columns:
                row[str(col_name)] = int(ct.loc[idx, col_name])
            freq_rows.append(row)
        freq_columns = [''] + [str(c) for c in ct.columns]

        # 卡方检验结果
        test_rows = [{
            '检验方法': '皮尔逊卡方检验',
            '卡方值(χ²)': round(float(chi2), 4),
            '自由度(df)': int(dof),
            'P值': round(float(p), 4),
            '结论': '显著 (p<0.05)' if p < 0.05 else '不显著 (p≥0.05)'
        }]
        sig = '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))

        # 堆叠柱状图
        ct_no_margin = pd.crosstab(df[row_var], df[col_var])
        chart = {
            'title': {'text': f'{row_var} × {col_var} 交叉分析', 'left': 'center'},
            'tooltip': {'trigger': 'axis', 'axisPointer': {'type': 'shadow'}},
            'legend': {'bottom': 0, 'data': [str(c) for c in ct_no_margin.columns]},
            'xAxis': {'type': 'category', 'data': [str(i) for i in ct_no_margin.index]},
            'yAxis': {'type': 'value', 'name': '频次'},
            'series': [{'name': str(c), 'type': 'bar', 'stack': 'total',
                        'data': [int(ct_no_margin.loc[i, c]) for i in ct_no_margin.index]}
                       for c in ct_no_margin.columns],
            'grid': {'bottom': 60}
        }
        return {
            'method_name': '交叉卡方分析',
            'summary_table': freq_rows,
            'columns': freq_columns,
            'extra_tables': [{'title': '卡方检验结果', 'rows': test_rows,
                              'columns': ['检验方法', '卡方值(χ²)', '自由度(df)', 'P值', '结论']}],
            'footnotes': [f'χ²={round(chi2,4)}, df={dof}, P={round(p,4)}{sig}',
                          '* p<0.05, ** p<0.01, *** p<0.001'],
            'charts': [chart],
            'interpretation': f'{row_var}与{col_var}的卡方检验 χ²={round(chi2,4)}，P={round(p,4)}，'
                              f'{"存在" if p < 0.05 else "不存在"}显著关联。'
        }


# ==============================================================================
#  ComparisonAnalyzer — 模块 3: 差异比较
# ==============================================================================

class ComparisonAnalyzer:
    """差异比较分析：参数/非参数检验"""

    @staticmethod
    def independent_ttest(df, value_col, group_col):
        """独立样本 t 检验"""
        tmp = df[[value_col, group_col]].dropna()
        tmp[value_col] = pd.to_numeric(tmp[value_col], errors='coerce')
        tmp = tmp.dropna()
        groups = tmp[group_col].unique()
        if len(groups) != 2:
            return {'method_name': '独立样本t检验', 'summary_table': [], 'columns': [],
                    'footnotes': [f'分组变量 {group_col} 应恰好有2个水平，当前有 {len(groups)} 个'],
                    'charts': [], 'interpretation': '分组变量水平数不等于2，无法执行独立样本t检验。', 'extra_tables': []}
        g1 = tmp[tmp[group_col] == groups[0]][value_col].values.astype(float)
        g2 = tmp[tmp[group_col] == groups[1]][value_col].values.astype(float)
        n1, n2 = len(g1), len(g2)
        m1, m2 = float(np.mean(g1)), float(np.mean(g2))
        sd1, sd2 = float(np.std(g1, ddof=1)), float(np.std(g2, ddof=1))
        se1, se2 = sd1 / np.sqrt(n1), sd2 / np.sqrt(n2)

        # Levene test
        lev_stat, lev_p = sp_stats.levene(g1, g2)
        equal_var = lev_p >= 0.05

        # t-test both ways
        t_eq, p_eq = sp_stats.ttest_ind(g1, g2, equal_var=True)
        t_we, p_we = sp_stats.ttest_ind(g1, g2, equal_var=False)
        df_eq = n1 + n2 - 2
        df_we = ((sd1**2/n1 + sd2**2/n2)**2 /
                 ((sd1**2/n1)**2/(n1-1) + (sd2**2/n2)**2/(n2-1)))

        # Cohen's d (pooled)
        pooled_std = np.sqrt(((n1-1)*sd1**2 + (n2-1)*sd2**2) / (n1+n2-2))
        cohen_d = (m1 - m2) / pooled_std if pooled_std > 0 else 0

        sig = lambda p: '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))
        chosen_t, chosen_p, chosen_df = (t_eq, p_eq, df_eq) if equal_var else (t_we, p_we, df_we)

        summary_table = [
            {'组别': str(groups[0]), 'N': n1, '均值': round(m1, 4), '标准差': round(sd1, 4), '标准误': round(se1, 4)},
            {'组别': str(groups[1]), 'N': n2, '均值': round(m2, 4), '标准差': round(sd2, 4), '标准误': round(se2, 4)},
        ]
        extra_tables = [
            {'title': 'Levene 方差齐性检验', 'columns': ['F值', 'P值', '结论'],
             'rows': [{'F值': round(float(lev_stat), 4), 'P值': round(float(lev_p), 4),
                       '结论': '方差齐' if equal_var else '方差不齐'}]},
            {'title': '独立样本t检验', 'columns': ['方法', 't值', '自由度(df)', 'P值(双侧)', '显著性', '均值差', 'Cohen\'s d'],
             'rows': [
                 {'方法': '假定方差相等', 't值': round(float(t_eq), 4), '自由度(df)': round(float(df_eq), 2),
                  'P值(双侧)': round(float(p_eq), 4), '显著性': sig(p_eq),
                  '均值差': round(m1 - m2, 4), 'Cohen\'s d': round(float(cohen_d), 4)},
                 {'方法': '不假定方差相等(Welch)', 't值': round(float(t_we), 4), '自由度(df)': round(float(df_we), 2),
                  'P值(双侧)': round(float(p_we), 4), '显著性': sig(p_we),
                  '均值差': round(m1 - m2, 4), 'Cohen\'s d': round(float(cohen_d), 4)},
             ]},
        ]
        method_label = '假定方差相等' if equal_var else '不假定方差相等(Welch)'
        interpretation = (f'Levene 检验 F={round(float(lev_stat),4)}, p={round(float(lev_p),4)}，'
                          f'{"方差齐" if equal_var else "方差不齐"}，采用{method_label}结果。'
                          f't={round(float(chosen_t),4)}, df={round(float(chosen_df),2)}, '
                          f'p={round(float(chosen_p),4)}{sig(chosen_p)}，Cohen\'s d={round(float(cohen_d),4)}。'
                          f'两组在 {value_col} 上的差异{"有" if chosen_p < 0.05 else "无"}统计学意义。')

        # boxplot
        chart = {
            'title': {'text': f'{value_col} 分组箱线图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': [str(groups[0]), str(groups[1])]},
            'yAxis': {'type': 'value', 'name': value_col},
            'series': [{
                'type': 'boxplot',
                'data': [
                    [round(float(v), 4) for v in [np.min(g1), np.percentile(g1,25), np.median(g1), np.percentile(g1,75), np.max(g1)]],
                    [round(float(v), 4) for v in [np.min(g2), np.percentile(g2,25), np.median(g2), np.percentile(g2,75), np.max(g2)]],
                ]
            }],
            'grid': {'bottom': 60}
        }
        return {'method_name': '独立样本t检验', 'summary_table': summary_table,
                'columns': ['组别', 'N', '均值', '标准差', '标准误'],
                'footnotes': [f'推荐方法: {method_label}', '* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}

    @staticmethod
    def paired_ttest(df, col1, col2):
        """配对样本 t 检验"""
        tmp = df[[col1, col2]].dropna()
        s1 = pd.to_numeric(tmp[col1], errors='coerce')
        s2 = pd.to_numeric(tmp[col2], errors='coerce')
        tmp = tmp.assign(_s1=s1, _s2=s2).dropna(subset=['_s1', '_s2'])
        s1 = tmp['_s1'].values.astype(float)
        s2 = tmp['_s2'].values.astype(float)
        if len(s1) < 2:
            return {'method_name': '配对t检验', 'summary_table': [], 'columns': [],
                    'footnotes': ['有效配对数不足'], 'charts': [], 'interpretation': '配对数据不足，无法分析。', 'extra_tables': []}
        diff = s1 - s2
        n = len(s1)
        t_stat, p_val = sp_stats.ttest_rel(s1, s2)
        mean_d, sd_d = float(np.mean(diff)), float(np.std(diff, ddof=1))
        cohen_d = mean_d / sd_d if sd_d > 0 else 0
        sig = '***' if p_val < 0.001 else ('**' if p_val < 0.01 else ('*' if p_val < 0.05 else ''))

        summary_table = [
            {'变量': col1, 'N': n, '均值': round(float(np.mean(s1)), 4), '标准差': round(float(np.std(s1, ddof=1)), 4), '标准误': round(float(np.std(s1, ddof=1)/np.sqrt(n)), 4)},
            {'变量': col2, 'N': n, '均值': round(float(np.mean(s2)), 4), '标准差': round(float(np.std(s2, ddof=1)), 4), '标准误': round(float(np.std(s2, ddof=1)/np.sqrt(n)), 4)},
            {'变量': '差值(1-2)', 'N': n, '均值': round(mean_d, 4), '标准差': round(sd_d, 4), '标准误': round(sd_d/np.sqrt(n), 4)},
        ]
        extra_tables = [
            {'title': '配对样本t检验', 'columns': ['t值', '自由度(df)', 'P值(双侧)', '显著性', '均值差', '差值标准差', 'Cohen\'s d'],
             'rows': [{'t值': round(float(t_stat), 4), '自由度(df)': n-1,
                       'P值(双侧)': round(float(p_val), 4), '显著性': sig,
                       '均值差': round(mean_d, 4), '差值标准差': round(sd_d, 4),
                       'Cohen\'s d': round(float(cohen_d), 4)}]}
        ]
        interpretation = (f'配对t检验: t={round(float(t_stat),4)}, df={n-1}, p={round(float(p_val),4)}{sig}，'
                          f'Cohen\'s d={round(float(cohen_d),4)}。'
                          f'{col1} 与 {col2} 之间的差异{"有" if p_val < 0.05 else "无"}统计学意义。')
        chart = {
            'title': {'text': f'{col1} vs {col2} 箱线图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': [col1, col2]},
            'yAxis': {'type': 'value'},
            'series': [{'type': 'boxplot', 'data': [
                [round(float(v), 4) for v in [np.min(s1), np.percentile(s1,25), np.median(s1), np.percentile(s1,75), np.max(s1)]],
                [round(float(v), 4) for v in [np.min(s2), np.percentile(s2,25), np.median(s2), np.percentile(s2,75), np.max(s2)]],
            ]}],
            'grid': {'bottom': 60}
        }
        return {'method_name': '配对t检验', 'summary_table': summary_table,
                'columns': ['变量', 'N', '均值', '标准差', '标准误'],
                'footnotes': ['* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}

    @staticmethod
    def one_way_anova(df, value_col, group_col, post_hoc=None):
        """单因素方差分析"""
        tmp = df[[value_col, group_col]].dropna()
        tmp[value_col] = pd.to_numeric(tmp[value_col], errors='coerce')
        tmp = tmp.dropna()
        group_names = sorted(tmp[group_col].unique(), key=str)
        k = len(group_names)
        if k < 2:
            return {'method_name': '单因素方差分析', 'summary_table': [], 'columns': [],
                    'footnotes': ['分组数不足2'], 'charts': [], 'interpretation': '分组变量水平数不足，无法分析。', 'extra_tables': []}

        groups = [tmp[tmp[group_col] == g][value_col].values.astype(float) for g in group_names]
        f_stat, p_val = sp_stats.f_oneway(*groups)

        # SS calculation
        grand_mean = float(np.mean(tmp[value_col]))
        N = len(tmp)
        ss_between = sum(len(g) * (np.mean(g) - grand_mean)**2 for g in groups)
        ss_within = sum(np.sum((g - np.mean(g))**2) for g in groups)
        ss_total = ss_between + ss_within
        df_between = k - 1
        df_within = N - k
        ms_between = ss_between / df_between if df_between > 0 else 0
        ms_within = ss_within / df_within if df_within > 0 else 0
        eta_sq = ss_between / ss_total if ss_total > 0 else 0

        sig_fn = lambda p: '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))

        summary_table = []
        box_data = []
        for i, g_name in enumerate(group_names):
            g = groups[i]
            n_g = len(g)
            m_g = float(np.mean(g))
            sd_g = float(np.std(g, ddof=1))
            se_g = sd_g / np.sqrt(n_g)
            ci_low = m_g - 1.96 * se_g
            ci_high = m_g + 1.96 * se_g
            summary_table.append({
                '组别': str(g_name), 'N': n_g, '均值': round(m_g, 4), '标准差': round(sd_g, 4),
                '标准误': round(se_g, 4), '95%CI下限': round(ci_low, 4), '95%CI上限': round(ci_high, 4)
            })
            box_data.append([round(float(v), 4) for v in [np.min(g), np.percentile(g,25), np.median(g), np.percentile(g,75), np.max(g)]])

        anova_table = [
            {'变异来源': '组间', 'SS': round(float(ss_between), 4), 'df': df_between,
             'MS': round(float(ms_between), 4), 'F': round(float(f_stat), 4),
             'P值': round(float(p_val), 4), '显著性': sig_fn(p_val), 'η²': round(float(eta_sq), 4)},
            {'变异来源': '组内', 'SS': round(float(ss_within), 4), 'df': df_within,
             'MS': round(float(ms_within), 4), 'F': '', 'P值': '', '显著性': '', 'η²': ''},
            {'变异来源': '总计', 'SS': round(float(ss_total), 4), 'df': N-1,
             'MS': '', 'F': '', 'P值': '', '显著性': '', 'η²': ''},
        ]
        extra_tables = [
            {'title': '方差分析表', 'columns': ['变异来源', 'SS', 'df', 'MS', 'F', 'P值', '显著性', 'η²'],
             'rows': anova_table}
        ]

        # Post-hoc tests
        if post_hoc and p_val < 0.05 and k >= 3:
            m = k * (k - 1) // 2
            posthoc_rows = []
            for i in range(k):
                for j in range(i+1, k):
                    gi, gj = groups[i], groups[j]
                    if post_hoc == 'tukey' and hasattr(sp_stats, 'tukey_hsd'):
                        # use tukey_hsd if available (scipy >= 1.8)
                        pass  # handled below
                    t_ph, p_ph = sp_stats.ttest_ind(gi, gj, equal_var=True)
                    p_adj = min(float(p_ph) * m, 1.0)
                    mean_diff = float(np.mean(gi) - np.mean(gj))
                    se_diff = np.sqrt(np.var(gi, ddof=1)/len(gi) + np.var(gj, ddof=1)/len(gj))
                    ci_low_ph = mean_diff - 1.96 * se_diff
                    ci_high_ph = mean_diff + 1.96 * se_diff
                    posthoc_rows.append({
                        '比较': f'{group_names[i]} vs {group_names[j]}',
                        '均值差': round(mean_diff, 4), '标准误': round(float(se_diff), 4),
                        'P值(校正)': round(p_adj, 4), '显著性': sig_fn(p_adj),
                        '95%CI下限': round(float(ci_low_ph), 4), '95%CI上限': round(float(ci_high_ph), 4),
                    })
            label = 'Bonferroni' if post_hoc == 'bonferroni' else 'Tukey(Bonferroni近似)'
            extra_tables.append({
                'title': f'事后多重比较 ({label})',
                'columns': ['比较', '均值差', '标准误', 'P值(校正)', '显著性', '95%CI下限', '95%CI上限'],
                'rows': posthoc_rows
            })

        interpretation = (f'单因素方差分析: F({df_between},{df_within})={round(float(f_stat),4)}, '
                          f'p={round(float(p_val),4)}{sig_fn(p_val)}, η²={round(float(eta_sq),4)}。'
                          f'各组在 {value_col} 上的差异{"有" if p_val < 0.05 else "无"}统计学意义。')

        chart = {
            'title': {'text': f'{value_col} 分组箱线图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': [str(g) for g in group_names]},
            'yAxis': {'type': 'value', 'name': value_col},
            'series': [{'type': 'boxplot', 'data': box_data}],
            'grid': {'bottom': 60}
        }
        return {'method_name': '单因素方差分析', 'summary_table': summary_table,
                'columns': ['组别', 'N', '均值', '标准差', '标准误', '95%CI下限', '95%CI上限'],
                'footnotes': ['* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}

    @staticmethod
    def mann_whitney(df, value_col, group_col):
        """Mann-Whitney U 检验"""
        tmp = df[[value_col, group_col]].dropna()
        tmp[value_col] = pd.to_numeric(tmp[value_col], errors='coerce')
        tmp = tmp.dropna()
        groups_u = tmp[group_col].unique()
        if len(groups_u) != 2:
            return {'method_name': 'Mann-Whitney U检验', 'summary_table': [], 'columns': [],
                    'footnotes': [f'分组变量应有2个水平，当前 {len(groups_u)} 个'],
                    'charts': [], 'interpretation': '分组变量水平数不等于2。', 'extra_tables': []}
        g1 = tmp[tmp[group_col] == groups_u[0]][value_col].values.astype(float)
        g2 = tmp[tmp[group_col] == groups_u[1]][value_col].values.astype(float)
        n1, n2 = len(g1), len(g2)

        u_stat, p_val = sp_stats.mannwhitneyu(g1, g2, alternative='two-sided')

        # Rank computation
        all_vals = pd.Series(np.concatenate([g1, g2]))
        ranks = all_vals.rank()
        rank1 = ranks.iloc[:n1]
        rank2 = ranks.iloc[n1:]
        rank_sum1 = float(rank1.sum())
        rank_sum2 = float(rank2.sum())
        mean_rank1 = rank_sum1 / n1
        mean_rank2 = rank_sum2 / n2

        # Z approximation
        mu = n1 * n2 / 2
        sigma = np.sqrt(n1 * n2 * (n1 + n2 + 1) / 12)
        z_val = (u_stat - mu) / sigma if sigma > 0 else 0
        r_effect = abs(z_val) / np.sqrt(n1 + n2)

        sig = '***' if p_val < 0.001 else ('**' if p_val < 0.01 else ('*' if p_val < 0.05 else ''))

        summary_table = [
            {'组别': str(groups_u[0]), 'N': n1, '中位数': round(float(np.median(g1)), 4),
             '均秩': round(mean_rank1, 4), '秩和': round(rank_sum1, 4)},
            {'组别': str(groups_u[1]), 'N': n2, '中位数': round(float(np.median(g2)), 4),
             '均秩': round(mean_rank2, 4), '秩和': round(rank_sum2, 4)},
        ]
        extra_tables = [
            {'title': 'Mann-Whitney U 检验', 'columns': ['U值', 'Z值', 'P值(双侧)', '显著性', '效应量r'],
             'rows': [{'U值': round(float(u_stat), 4), 'Z值': round(float(z_val), 4),
                       'P值(双侧)': round(float(p_val), 4), '显著性': sig,
                       '效应量r': round(float(r_effect), 4)}]}
        ]
        interpretation = (f'Mann-Whitney U检验: U={round(float(u_stat),4)}, Z={round(float(z_val),4)}, '
                          f'p={round(float(p_val),4)}{sig}, r={round(float(r_effect),4)}。'
                          f'两组在 {value_col} 上的差异{"有" if p_val < 0.05 else "无"}统计学意义。')

        chart = {
            'title': {'text': f'{value_col} 分组箱线图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': [str(groups_u[0]), str(groups_u[1])]},
            'yAxis': {'type': 'value', 'name': value_col},
            'series': [{'type': 'boxplot', 'data': [
                [round(float(v), 4) for v in [np.min(g1), np.percentile(g1,25), np.median(g1), np.percentile(g1,75), np.max(g1)]],
                [round(float(v), 4) for v in [np.min(g2), np.percentile(g2,25), np.median(g2), np.percentile(g2,75), np.max(g2)]],
            ]}],
            'grid': {'bottom': 60}
        }
        return {'method_name': 'Mann-Whitney U检验', 'summary_table': summary_table,
                'columns': ['组别', 'N', '中位数', '均秩', '秩和'],
                'footnotes': ['* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}

    @staticmethod
    def wilcoxon_test(df, col1, col2):
        """Wilcoxon 符号秩检验"""
        tmp = df[[col1, col2]].dropna()
        s1 = pd.to_numeric(tmp[col1], errors='coerce')
        s2 = pd.to_numeric(tmp[col2], errors='coerce')
        tmp = tmp.assign(_s1=s1, _s2=s2).dropna(subset=['_s1', '_s2'])
        s1 = tmp['_s1'].values.astype(float)
        s2 = tmp['_s2'].values.astype(float)
        diff = s1 - s2
        non_zero = diff[diff != 0]
        n_total = len(s1)
        n_nz = len(non_zero)
        if n_nz < 2:
            return {'method_name': 'Wilcoxon符号秩检验', 'summary_table': [], 'columns': [],
                    'footnotes': ['有效非零差值对不足'], 'charts': [], 'interpretation': '数据不足。', 'extra_tables': []}

        w_stat, p_val = sp_stats.wilcoxon(s1, s2)
        # Z approximation
        mu_w = n_nz * (n_nz + 1) / 4
        sigma_w = np.sqrt(n_nz * (n_nz + 1) * (2*n_nz + 1) / 24)
        z_val = (w_stat - mu_w) / sigma_w if sigma_w > 0 else 0
        r_effect = abs(z_val) / np.sqrt(n_total)

        sig = '***' if p_val < 0.001 else ('**' if p_val < 0.01 else ('*' if p_val < 0.05 else ''))

        summary_table = [
            {'变量': col1, 'N': n_total, '中位数': round(float(np.median(s1)), 4),
             'Q1': round(float(np.percentile(s1, 25)), 4), 'Q3': round(float(np.percentile(s1, 75)), 4)},
            {'变量': col2, 'N': n_total, '中位数': round(float(np.median(s2)), 4),
             'Q1': round(float(np.percentile(s2, 25)), 4), 'Q3': round(float(np.percentile(s2, 75)), 4)},
        ]
        extra_tables = [
            {'title': 'Wilcoxon 符号秩检验', 'columns': ['T统计量', 'Z值', 'P值(双侧)', '显著性', '效应量r'],
             'rows': [{'T统计量': round(float(w_stat), 4), 'Z值': round(float(z_val), 4),
                       'P值(双侧)': round(float(p_val), 4), '显著性': sig,
                       '效应量r': round(float(r_effect), 4)}]}
        ]
        interpretation = (f'Wilcoxon符号秩检验: T={round(float(w_stat),4)}, Z={round(float(z_val),4)}, '
                          f'p={round(float(p_val),4)}{sig}, r={round(float(r_effect),4)}。'
                          f'{col1} 与 {col2} 之间的差异{"有" if p_val < 0.05 else "无"}统计学意义。')

        chart = {
            'title': {'text': f'{col1} vs {col2} 箱线图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': [col1, col2]},
            'yAxis': {'type': 'value'},
            'series': [{'type': 'boxplot', 'data': [
                [round(float(v), 4) for v in [np.min(s1), np.percentile(s1,25), np.median(s1), np.percentile(s1,75), np.max(s1)]],
                [round(float(v), 4) for v in [np.min(s2), np.percentile(s2,25), np.median(s2), np.percentile(s2,75), np.max(s2)]],
            ]}],
            'grid': {'bottom': 60}
        }
        return {'method_name': 'Wilcoxon符号秩检验', 'summary_table': summary_table,
                'columns': ['变量', 'N', '中位数', 'Q1', 'Q3'],
                'footnotes': ['* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}


# ==============================================================================
#  RegressionAnalyzer — 模块 4: 相关与回归
# ==============================================================================

class RegressionAnalyzer:
    """相关与回归分析"""

    @staticmethod
    def pearson_correlation(df, cols):
        """Pearson 相关分析"""
        sub = df[cols].apply(pd.to_numeric, errors='coerce').dropna()
        if len(sub) < 3:
            return {'method_name': 'Pearson相关分析', 'summary_table': [], 'columns': [],
                    'footnotes': ['样本量不足'], 'charts': [], 'interpretation': '数据不足。', 'extra_tables': []}
        n = len(cols)
        sig_fn = lambda p: '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))

        r_matrix = []
        p_matrix = []
        heatmap_data = []
        max_r, max_pair = 0, ''
        for i in range(n):
            r_row = {'变量': cols[i]}
            p_row = {'变量': cols[i]}
            for j in range(n):
                if i == j:
                    r_row[cols[j]] = '1'
                    p_row[cols[j]] = '-'
                    heatmap_data.append([i, j, 1.0])
                else:
                    r, p = sp_stats.pearsonr(sub[cols[i]].values, sub[cols[j]].values)
                    r_row[cols[j]] = f'{round(float(r),4)}{sig_fn(p)}'
                    p_row[cols[j]] = round(float(p), 4)
                    heatmap_data.append([i, j, round(float(r), 4)])
                    if i < j and abs(r) > max_r:
                        max_r = abs(r)
                        max_pair = f'{cols[i]} 与 {cols[j]}'
            r_matrix.append(r_row)
            p_matrix.append(p_row)

        chart = {
            'title': {'text': 'Pearson 相关系数热力图', 'left': 'center'},
            'tooltip': {'formatter': '{c}'},
            'xAxis': {'type': 'category', 'data': cols, 'axisLabel': {'rotate': len(cols) > 6 and 45 or 0, 'fontSize': 11}},
            'yAxis': {'type': 'category', 'data': cols},
            'visualMap': {'min': -1, 'max': 1, 'calculable': True,
                          'inRange': {'color': ['#2563eb', '#ffffff', '#dc2626']},
                          'orient': 'horizontal', 'left': 'center', 'bottom': 0},
            'series': [{'type': 'heatmap', 'data': heatmap_data,
                        'label': {'show': True, 'fontSize': 10},
                        'emphasis': {'itemStyle': {'shadowBlur': 10}}}],
            'grid': {'bottom': 80, 'top': 40}
        }
        interpretation = f'Pearson相关分析(N={len(sub)})。最强相关: {max_pair}(|r|={round(max_r,4)})。' if max_pair else f'Pearson相关分析(N={len(sub)})。'

        return {'method_name': 'Pearson相关分析', 'summary_table': r_matrix,
                'columns': ['变量'] + cols,
                'footnotes': [f'N={len(sub)}', '* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation,
                'extra_tables': [{'title': 'P值矩阵', 'columns': ['变量'] + cols, 'rows': p_matrix}]}

    @staticmethod
    def spearman_correlation(df, cols):
        """Spearman 秩相关分析"""
        sub = df[cols].apply(pd.to_numeric, errors='coerce').dropna()
        if len(sub) < 3:
            return {'method_name': 'Spearman秩相关分析', 'summary_table': [], 'columns': [],
                    'footnotes': ['样本量不足'], 'charts': [], 'interpretation': '数据不足。', 'extra_tables': []}
        n = len(cols)
        sig_fn = lambda p: '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))

        r_matrix = []
        p_matrix = []
        heatmap_data = []
        max_r, max_pair = 0, ''
        for i in range(n):
            r_row = {'变量': cols[i]}
            p_row = {'变量': cols[i]}
            for j in range(n):
                if i == j:
                    r_row[cols[j]] = '1'
                    p_row[cols[j]] = '-'
                    heatmap_data.append([i, j, 1.0])
                else:
                    r, p = sp_stats.spearmanr(sub[cols[i]].values, sub[cols[j]].values)
                    r_row[cols[j]] = f'{round(float(r),4)}{sig_fn(p)}'
                    p_row[cols[j]] = round(float(p), 4)
                    heatmap_data.append([i, j, round(float(r), 4)])
                    if i < j and abs(r) > max_r:
                        max_r = abs(r)
                        max_pair = f'{cols[i]} 与 {cols[j]}'
            r_matrix.append(r_row)
            p_matrix.append(p_row)

        chart = {
            'title': {'text': 'Spearman 秩相关系数热力图', 'left': 'center'},
            'tooltip': {'formatter': '{c}'},
            'xAxis': {'type': 'category', 'data': cols, 'axisLabel': {'rotate': len(cols) > 6 and 45 or 0, 'fontSize': 11}},
            'yAxis': {'type': 'category', 'data': cols},
            'visualMap': {'min': -1, 'max': 1, 'calculable': True,
                          'inRange': {'color': ['#2563eb', '#ffffff', '#dc2626']},
                          'orient': 'horizontal', 'left': 'center', 'bottom': 0},
            'series': [{'type': 'heatmap', 'data': heatmap_data,
                        'label': {'show': True, 'fontSize': 10},
                        'emphasis': {'itemStyle': {'shadowBlur': 10}}}],
            'grid': {'bottom': 80, 'top': 40}
        }
        interpretation = f'Spearman秩相关分析(N={len(sub)})。最强相关: {max_pair}(|ρ|={round(max_r,4)})。' if max_pair else f'Spearman秩相关分析(N={len(sub)})。'

        return {'method_name': 'Spearman秩相关分析', 'summary_table': r_matrix,
                'columns': ['变量'] + cols,
                'footnotes': [f'N={len(sub)}', '* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation,
                'extra_tables': [{'title': 'P值矩阵', 'columns': ['变量'] + cols, 'rows': p_matrix}]}

    @staticmethod
    def linear_regression(df, y_col, x_cols):
        """多元线性回归 (OLS)"""
        all_cols = [y_col] + list(x_cols)
        sub = df[all_cols].apply(pd.to_numeric, errors='coerce').dropna()
        n = len(sub)
        k = len(x_cols)
        if n < k + 2:
            return {'method_name': '线性回归分析', 'summary_table': [], 'columns': [],
                    'footnotes': ['样本量不足'], 'charts': [], 'interpretation': '数据不足。', 'extra_tables': []}

        y = sub[y_col].values.astype(float)
        X_raw = sub[x_cols].values.astype(float)
        X = np.column_stack([np.ones(n), X_raw])  # add intercept

        # OLS: beta = (X'X)^-1 X'y
        try:
            XtX_inv = np.linalg.inv(X.T @ X)
        except np.linalg.LinAlgError:
            return {'method_name': '线性回归分析', 'summary_table': [], 'columns': [],
                    'footnotes': ['设计矩阵奇异，无法求解'], 'charts': [],
                    'interpretation': '自变量间可能存在完全多重共线性，无法求解。', 'extra_tables': []}

        beta = XtX_inv @ X.T @ y
        y_hat = X @ beta
        residuals = y - y_hat
        y_mean = float(np.mean(y))

        SSE = float(np.sum(residuals**2))
        SST = float(np.sum((y - y_mean)**2))
        SSR = SST - SSE
        R2 = 1 - SSE / SST if SST > 0 else 0
        adj_R2 = 1 - (1 - R2) * (n - 1) / (n - k - 1) if n - k - 1 > 0 else 0
        MSE = SSE / (n - k - 1) if n - k - 1 > 0 else 0
        MSR = SSR / k if k > 0 else 0
        F_stat = MSR / MSE if MSE > 0 else 0
        p_F = float(sp_stats.f.sf(F_stat, k, n - k - 1)) if n - k - 1 > 0 else 1.0
        se_est = np.sqrt(MSE)

        # Coefficient SE, t, p, CI
        se_beta = np.sqrt(np.diag(XtX_inv) * MSE)
        # Standardized Beta
        std_y = float(np.std(y, ddof=1))
        std_x = np.std(X_raw, axis=0, ddof=1)

        sig_fn = lambda p: '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))
        t_crit = float(sp_stats.t.ppf(0.975, n - k - 1)) if n - k - 1 > 0 else 1.96

        var_names = ['(常量)'] + list(x_cols)
        summary_table = []
        for i, vname in enumerate(var_names):
            b = float(beta[i])
            se = float(se_beta[i])
            t_val = b / se if se > 0 else 0
            p_val = 2 * float(sp_stats.t.sf(abs(t_val), n - k - 1)) if n - k - 1 > 0 else 1.0
            ci_low = b - t_crit * se
            ci_high = b + t_crit * se
            if i == 0:
                beta_std = '-'
            else:
                beta_std = round(b * float(std_x[i-1]) / std_y, 4) if std_y > 0 else '-'
            summary_table.append({
                '变量': vname, '非标准化系数B': round(b, 4), '标准误': round(se, 4),
                'Beta(标准化)': beta_std, 't值': round(float(t_val), 4),
                'P值': round(float(p_val), 4), '显著性': sig_fn(p_val),
                '95%CI下限': round(float(ci_low), 4), '95%CI上限': round(float(ci_high), 4),
            })

        extra_tables = [
            {'title': '模型摘要', 'columns': ['R', 'R²', '调整R²', '标准估计误差', 'F值', 'df1', 'df2', 'P值(F)', '显著性'],
             'rows': [{'R': round(float(np.sqrt(R2)), 4), 'R²': round(float(R2), 4),
                       '调整R²': round(float(adj_R2), 4), '标准估计误差': round(float(se_est), 4),
                       'F值': round(float(F_stat), 4), 'df1': k, 'df2': n-k-1,
                       'P值(F)': round(float(p_F), 4), '显著性': sig_fn(p_F)}]}
        ]

        interpretation = (f'线性回归模型: R²={round(float(R2),4)}, 调整R²={round(float(adj_R2),4)}, '
                          f'F({k},{n-k-1})={round(float(F_stat),4)}, p={round(float(p_F),4)}{sig_fn(p_F)}。'
                          f'模型{"有" if p_F < 0.05 else "无"}统计学意义。')

        # Chart
        if k == 1:
            # scatter + regression line
            x_vals = X_raw[:, 0]
            x_sorted = np.sort(x_vals)
            y_line = float(beta[0]) + float(beta[1]) * x_sorted
            chart = {
                'title': {'text': f'{x_cols[0]} → {y_col} 回归散点图', 'left': 'center'},
                'tooltip': {'trigger': 'item'},
                'xAxis': {'type': 'value', 'name': x_cols[0]},
                'yAxis': {'type': 'value', 'name': y_col},
                'series': [
                    {'type': 'scatter', 'data': [[round(float(x_vals[i]),4), round(float(y[i]),4)] for i in range(n)],
                     'itemStyle': {'color': '#5b8ff9', 'opacity': 0.6}},
                    {'type': 'line', 'data': [[round(float(x_sorted[i]),4), round(float(y_line[i]),4)] for i in range(len(x_sorted))],
                     'lineStyle': {'color': '#dc2626', 'width': 2}, 'symbol': 'none', 'name': '回归线'},
                ],
                'grid': {'bottom': 60}
            }
        else:
            # residual plot
            chart = {
                'title': {'text': '残差图 (预测值 vs 残差)', 'left': 'center'},
                'tooltip': {'trigger': 'item'},
                'xAxis': {'type': 'value', 'name': '预测值'},
                'yAxis': {'type': 'value', 'name': '残差'},
                'series': [
                    {'type': 'scatter', 'data': [[round(float(y_hat[i]),4), round(float(residuals[i]),4)] for i in range(n)],
                     'itemStyle': {'color': '#5b8ff9', 'opacity': 0.6}},
                ],
                'grid': {'bottom': 60}
            }
        return {'method_name': '线性回归分析', 'summary_table': summary_table,
                'columns': ['变量', '非标准化系数B', '标准误', 'Beta(标准化)', 't值', 'P值', '显著性', '95%CI下限', '95%CI上限'],
                'footnotes': [f'N={n}, 自变量数={k}', '* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}

    @staticmethod
    def logistic_regression(df, y_col, x_cols, max_iter=100):
        """Logistic 回归（二分类）"""
        from scipy.optimize import minimize

        all_cols = [y_col] + list(x_cols)
        sub = df[all_cols].dropna()
        # Encode y to 0/1
        y_raw = sub[y_col]
        unique_vals = sorted(y_raw.unique(), key=str)
        if len(unique_vals) != 2:
            return {'method_name': 'Logistic回归', 'summary_table': [], 'columns': [],
                    'footnotes': [f'因变量应为二分类，当前有 {len(unique_vals)} 个值'],
                    'charts': [], 'interpretation': '因变量不是二分类变量。', 'extra_tables': []}
        # Map to 0/1
        val_map = {unique_vals[0]: 0, unique_vals[1]: 1}
        y = np.array([val_map[v] for v in y_raw.values], dtype=float)

        X_raw = sub[x_cols].apply(pd.to_numeric, errors='coerce').values.astype(float)
        # Check for NaN after numeric conversion
        valid_mask = ~np.any(np.isnan(X_raw), axis=1)
        X_raw = X_raw[valid_mask]
        y = y[valid_mask]
        n = len(y)
        k = len(x_cols)
        if n < k + 2:
            return {'method_name': 'Logistic回归', 'summary_table': [], 'columns': [],
                    'footnotes': ['样本量不足'], 'charts': [], 'interpretation': '数据不足。', 'extra_tables': []}

        # Standardize X for numerical stability
        x_mean = np.mean(X_raw, axis=0)
        x_std = np.std(X_raw, axis=0, ddof=1)
        x_std[x_std == 0] = 1
        X_scaled = (X_raw - x_mean) / x_std
        X = np.column_stack([np.ones(n), X_scaled])

        eps = 1e-15

        def neg_log_likelihood(beta):
            z = np.clip(X @ beta, -500, 500)
            p = 1.0 / (1.0 + np.exp(-z))
            return -np.sum(y * np.log(p + eps) + (1 - y) * np.log(1 - p + eps))

        def gradient(beta):
            z = np.clip(X @ beta, -500, 500)
            p = 1.0 / (1.0 + np.exp(-z))
            return X.T @ (p - y)

        beta0 = np.zeros(k + 1)
        result = minimize(neg_log_likelihood, beta0, jac=gradient, method='L-BFGS-B',
                          options={'maxiter': int(max_iter), 'disp': False})
        beta_scaled = result.x
        converged = result.success
        ll_model = -result.fun

        # Null model log-likelihood
        p_null = np.mean(y)
        ll_null = float(np.sum(y * np.log(p_null + eps) + (1 - y) * np.log(1 - p_null + eps)))

        # Convert back to original scale
        beta_orig = np.zeros(k + 1)
        beta_orig[0] = beta_scaled[0] - np.sum(beta_scaled[1:] * x_mean / x_std)
        beta_orig[1:] = beta_scaled[1:] / x_std

        # SE via Fisher information matrix (original scale)
        X_orig = np.column_stack([np.ones(n), X_raw])
        z_orig = np.clip(X_orig @ beta_orig, -500, 500)
        p_hat = 1.0 / (1.0 + np.exp(-z_orig))
        W = p_hat * (1 - p_hat)
        try:
            fisher_info = X_orig.T @ np.diag(W) @ X_orig
            cov_matrix = np.linalg.inv(fisher_info)
            se_beta = np.sqrt(np.abs(np.diag(cov_matrix)))
        except np.linalg.LinAlgError:
            se_beta = np.full(k + 1, np.nan)

        sig_fn = lambda p: '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))

        var_names = ['(常量)'] + list(x_cols)
        summary_table = []
        or_values = []
        or_lower = []
        or_upper = []
        for i, vname in enumerate(var_names):
            b = float(beta_orig[i])
            se = float(se_beta[i])
            wald = (b / se)**2 if se > 0 and not np.isnan(se) else np.nan
            p_val = float(sp_stats.chi2.sf(wald, 1)) if not np.isnan(wald) else np.nan
            or_val = float(np.exp(np.clip(b, -500, 500)))
            ci_low = float(np.exp(np.clip(b - 1.96 * se, -500, 500))) if not np.isnan(se) else np.nan
            ci_high = float(np.exp(np.clip(b + 1.96 * se, -500, 500))) if not np.isnan(se) else np.nan

            p_display = round(float(p_val), 4) if not np.isnan(p_val) else '-'
            sig_display = sig_fn(p_val) if not np.isnan(p_val) else ''

            summary_table.append({
                '变量': vname, 'B': round(b, 4),
                '标准误': round(se, 4) if not np.isnan(se) else '-',
                'Wald χ²': round(float(wald), 4) if not np.isnan(wald) else '-',
                'df': 1, 'P值': p_display, '显著性': sig_display,
                'OR(Exp(B))': round(or_val, 4),
                'OR 95%CI下限': round(ci_low, 4) if not np.isnan(ci_low) else '-',
                'OR 95%CI上限': round(ci_high, 4) if not np.isnan(ci_high) else '-',
            })
            if i > 0:
                or_values.append(or_val)
                or_lower.append(ci_low if not np.isnan(ci_low) else or_val)
                or_upper.append(ci_high if not np.isnan(ci_high) else or_val)

        # Model summary
        neg2ll = -2 * ll_model
        cox_snell = 1 - np.exp(2 * (ll_null - ll_model) / n)
        nagelkerke = cox_snell / (1 - np.exp(2 * ll_null / n)) if (1 - np.exp(2 * ll_null / n)) > 0 else 0
        n_event = int(np.sum(y))

        # Classification table
        y_pred = (p_hat >= 0.5).astype(int)
        tp = int(np.sum((y == 1) & (y_pred == 1)))
        tn = int(np.sum((y == 0) & (y_pred == 0)))
        fp = int(np.sum((y == 0) & (y_pred == 1)))
        fn = int(np.sum((y == 1) & (y_pred == 0)))
        acc = (tp + tn) / n * 100 if n > 0 else 0

        extra_tables = [
            {'title': '模型摘要', 'columns': ['样本量', '事件数', '-2LL', 'Cox & Snell R²', 'Nagelkerke R²'],
             'rows': [{'样本量': n, '事件数': n_event,
                       '-2LL': round(float(neg2ll), 4),
                       'Cox & Snell R²': round(float(cox_snell), 4),
                       'Nagelkerke R²': round(float(nagelkerke), 4)}]},
            {'title': '分类表', 'columns': ['观测值', '预测=0', '预测=1', '正确率(%)'],
             'rows': [
                 {'观测值': f'0({unique_vals[0]})', '预测=0': tn, '预测=1': fp,
                  '正确率(%)': round(tn/(tn+fp)*100, 2) if tn+fp > 0 else 0},
                 {'观测值': f'1({unique_vals[1]})', '预测=0': fn, '预测=1': tp,
                  '正确率(%)': round(tp/(tp+fn)*100, 2) if tp+fn > 0 else 0},
                 {'观测值': '总体', '预测=0': tn+fn, '预测=1': fp+tp,
                  '正确率(%)': round(acc, 2)},
             ]},
        ]

        conv_note = '' if converged else '（注意：模型未收敛，结果可能不可靠）'
        interpretation = (f'Logistic回归{conv_note}: -2LL={round(float(neg2ll),4)}, '
                          f'Nagelkerke R²={round(float(nagelkerke),4)}, '
                          f'总体正确率={round(acc,2)}%。')

        # OR forest plot
        chart = {
            'title': {'text': 'OR值森林图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'value', 'name': 'OR', 'min': 0,
                      'axisLine': {'lineStyle': {'color': '#999'}}},
            'yAxis': {'type': 'category', 'data': list(x_cols),
                      'axisLabel': {'fontSize': 12}},
            'series': [
                {'type': 'scatter', 'data': [round(v, 4) for v in or_values],
                 'symbolSize': 10, 'itemStyle': {'color': '#2563eb'}},
                {'type': 'custom',
                 'renderItem': None,  # placeholder, ECharts needs proper custom rendering
                 'data': []},
            ],
            'grid': {'left': 120, 'right': 40, 'bottom': 40}
        }
        # Simplified: use error bar via markLine approach instead
        # Replace with a simpler bar-like representation
        chart = {
            'title': {'text': 'OR值及95%CI', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'category', 'data': list(x_cols),
                      'axisLabel': {'rotate': len(x_cols) > 4 and 30 or 0}},
            'yAxis': {'type': 'value', 'name': 'OR', 'min': 0},
            'series': [
                {'type': 'bar', 'data': [round(v, 4) for v in or_values],
                 'itemStyle': {'color': '#2563eb'}, 'barWidth': '40%', 'name': 'OR'},
                {'type': 'scatter', 'data': [{'value': round(or_lower[i], 4), 'symbolSize': 0} for i in range(k)],
                 'itemStyle': {'color': 'transparent'}, 'name': ''},
                {'type': 'scatter', 'data': [{'value': round(or_upper[i], 4), 'symbolSize': 0} for i in range(k)],
                 'itemStyle': {'color': 'transparent'}, 'name': ''},
            ],
            'grid': {'bottom': 60}
        }
        # Add markLine at OR=1
        chart['series'][0]['markLine'] = {
            'data': [{'yAxis': 1}],
            'lineStyle': {'color': '#dc2626', 'type': 'dashed'},
            'label': {'formatter': 'OR=1'},
            'symbol': 'none'
        }

        return {'method_name': 'Logistic回归', 'summary_table': summary_table,
                'columns': ['变量', 'B', '标准误', 'Wald χ²', 'df', 'P值', '显著性', 'OR(Exp(B))', 'OR 95%CI下限', 'OR 95%CI上限'],
                'footnotes': [f'N={n}, 事件数={n_event}', f'因变量: {y_col} (0={unique_vals[0]}, 1={unique_vals[1]})',
                              '* p<0.05, ** p<0.01, *** p<0.001'],
                'charts': [chart], 'interpretation': interpretation, 'extra_tables': extra_tables}


# ==============================================================================
#  PrerequisiteTests — 模块 5: 统计前提检验
# ==============================================================================

class PrerequisiteTests:

    @staticmethod
    def normality_test(df, cols):
        """正态性检验：n≤50 Shapiro-Wilk，n>50 K-S 检验"""
        rows = []
        for col in cols:
            s = pd.to_numeric(df[col], errors='coerce').dropna()
            n = len(s)
            if n < 3:
                rows.append({'变量': col, 'N': n, '检验方法': '-', '统计量': '-', 'P值': '-', '结论': '样本量不足'})
                continue
            if n <= 50:
                stat, p = sp_stats.shapiro(s)
                method = 'Shapiro-Wilk'
                stat_label = 'W'
            else:
                stat, p = sp_stats.kstest(s, 'norm', args=(s.mean(), s.std()))
                method = 'Kolmogorov-Smirnov'
                stat_label = 'D'
            sig = '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))
            rows.append({
                '变量': col, 'N': n, '检验方法': method,
                f'统计量({stat_label})': round(float(stat), 4),
                'P值': round(float(p), 4),
                '显著性': sig,
                '结论': '服从正态分布' if p >= 0.05 else '不服从正态分布'
            })

        columns = ['变量', 'N', '检验方法']
        if rows and any('统计量(W)' in r for r in rows):
            columns.append('统计量(W)')
        if rows and any('统计量(D)' in r for r in rows):
            columns.append('统计量(D)')
        columns += ['P值', '显著性', '结论']

        return {
            'method_name': '正态性检验',
            'summary_table': rows,
            'columns': columns,
            'footnotes': ['Shapiro-Wilk 检验适用于 n≤50；K-S 检验适用于 n>50',
                          '* p<0.05, ** p<0.01, *** p<0.001',
                          'P≥0.05 表示服从正态分布'],
            'charts': [],
            'interpretation': '。'.join([
                f'{r["变量"]}：{r["结论"]}' for r in rows if isinstance(r.get("结论"), str)
            ])
        }

    @staticmethod
    def homogeneity_test(df, value_col, group_col):
        """Levene 方差齐性检验"""
        groups = []
        group_names = []
        for name, gdf in df.groupby(group_col):
            s = pd.to_numeric(gdf[value_col], errors='coerce').dropna()
            if len(s) > 0:
                groups.append(s.values)
                group_names.append(str(name))
        if len(groups) < 2:
            return {
                'method_name': 'Levene 方差齐性检验',
                'summary_table': [{'变量': value_col, '分组变量': group_col,
                                   'msg': '分组数不足，至少需要 2 组'}],
                'columns': ['变量', '分组变量', 'msg'],
                'footnotes': [], 'charts': [],
                'interpretation': '分组数不足，无法执行检验。'
            }
        f_stat, p = sp_stats.levene(*groups)
        sig = '***' if p < 0.001 else ('**' if p < 0.01 else ('*' if p < 0.05 else ''))
        rows = [{
            '检验变量': value_col, '分组变量': group_col,
            '组数': len(groups),
            'Levene F': round(float(f_stat), 4),
            'df1': len(groups) - 1,
            'df2': sum(len(g) for g in groups) - len(groups),
            'P值': round(float(p), 4),
            '显著性': sig,
            '结论': '方差齐' if p >= 0.05 else '方差不齐'
        }]
        return {
            'method_name': 'Levene 方差齐性检验',
            'summary_table': rows,
            'columns': ['检验变量', '分组变量', '组数', 'Levene F', 'df1', 'df2', 'P值', '显著性', '结论'],
            'footnotes': ['P≥0.05 表示方差齐性假设成立', '* p<0.05, ** p<0.01, *** p<0.001'],
            'charts': [],
            'interpretation': f'{value_col}按{group_col}分组的 Levene 检验 F={round(f_stat,4)}，P={round(p,4)}，'
                              f'{"满足" if p >= 0.05 else "不满足"}方差齐性假设。'
        }


# ==============================================================================
#  ChartGenerator — 模块 6: 可视化
# ==============================================================================

class ChartGenerator:

    @staticmethod
    def histogram(df, col, bins=None, title=None):
        s = pd.to_numeric(df[col], errors='coerce').dropna()
        if len(s) == 0:
            return {'error': '无有效数值数据'}
        if bins is None:
            bins = min(30, max(5, int(np.sqrt(len(s)))))
        counts, edges = np.histogram(s, bins=bins)
        labels = [f'{round(edges[i],2)}-{round(edges[i+1],2)}' for i in range(len(counts))]
        return {
            'title': {'text': title or f'{col} 直方图', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'category', 'data': labels, 'axisLabel': {'rotate': 45, 'fontSize': 10}},
            'yAxis': {'type': 'value', 'name': '频次'},
            'series': [{'type': 'bar', 'data': [int(c) for c in counts],
                        'itemStyle': {'color': '#5b8ff9'}, 'barWidth': '90%'}],
            'grid': {'bottom': 100, 'left': 60}
        }

    @staticmethod
    def kde_plot(df, col, title=None):
        s = pd.to_numeric(df[col], errors='coerce').dropna()
        if len(s) < 2:
            return {'error': '样本量不足'}
        from scipy.stats import gaussian_kde
        kde = gaussian_kde(s)
        x_range = np.linspace(float(s.min()), float(s.max()), 200)
        y = kde(x_range)
        return {
            'title': {'text': title or f'{col} 核密度图', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'value', 'name': col},
            'yAxis': {'type': 'value', 'name': '密度'},
            'series': [{'type': 'line', 'data': [[round(float(x_range[i]), 4), round(float(y[i]), 6)]
                                                  for i in range(len(x_range))],
                        'smooth': True, 'areaStyle': {'opacity': 0.3},
                        'itemStyle': {'color': '#5b8ff9'}, 'showSymbol': False}],
            'grid': {'bottom': 60, 'left': 70}
        }

    @staticmethod
    def boxplot(df, cols, group_var=None, title=None):
        box_data = []
        labels = []
        if group_var and group_var in df.columns:
            for col in cols:
                for gname, gdf in df.groupby(group_var):
                    s = pd.to_numeric(gdf[col], errors='coerce').dropna()
                    if len(s) > 0:
                        box_data.append([float(s.min()), float(s.quantile(0.25)),
                                         float(s.median()), float(s.quantile(0.75)), float(s.max())])
                        labels.append(f'{col}({gname})')
        else:
            for col in cols:
                s = pd.to_numeric(df[col], errors='coerce').dropna()
                if len(s) > 0:
                    box_data.append([float(s.min()), float(s.quantile(0.25)),
                                     float(s.median()), float(s.quantile(0.75)), float(s.max())])
                    labels.append(col)
        return {
            'title': {'text': title or '箱线图', 'left': 'center'},
            'tooltip': {'trigger': 'item'},
            'xAxis': {'type': 'category', 'data': labels, 'axisLabel': {'rotate': 30}},
            'yAxis': {'type': 'value'},
            'series': [{'type': 'boxplot', 'data': box_data,
                        'itemStyle': {'color': '#dbeafe', 'borderColor': '#2563eb'}}],
            'grid': {'bottom': 80}
        }

    @staticmethod
    def bar_chart(df, col, title=None, horizontal=False):
        vc = df[col].value_counts().sort_index()
        labels = [str(v) for v in vc.index]
        values = [int(v) for v in vc.values]
        if horizontal:
            return {
                'title': {'text': title or f'{col} 条形图', 'left': 'center'},
                'tooltip': {'trigger': 'axis'},
                'yAxis': {'type': 'category', 'data': labels},
                'xAxis': {'type': 'value', 'name': '频次'},
                'series': [{'type': 'bar', 'data': values, 'itemStyle': {'color': '#5b8ff9'},
                            'label': {'show': True, 'position': 'right'}}],
                'grid': {'left': 120, 'bottom': 40}
            }
        return {
            'title': {'text': title or f'{col} 柱状图', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'category', 'data': labels, 'axisLabel': {'rotate': 30}},
            'yAxis': {'type': 'value', 'name': '频次'},
            'series': [{'type': 'bar', 'data': values, 'itemStyle': {'color': '#5b8ff9'},
                        'label': {'show': True, 'position': 'top'}}],
            'grid': {'bottom': 80}
        }

    @staticmethod
    def scatter_plot(df, x_col, y_col, title=None):
        x = pd.to_numeric(df[x_col], errors='coerce')
        y = pd.to_numeric(df[y_col], errors='coerce')
        valid = x.notna() & y.notna()
        data = [[round(float(x.iloc[i]), 4), round(float(y.iloc[i]), 4)]
                for i in range(len(x)) if valid.iloc[i]]
        return {
            'title': {'text': title or f'{x_col} vs {y_col}', 'left': 'center'},
            'tooltip': {'trigger': 'item', 'formatter': '{c}'},
            'xAxis': {'type': 'value', 'name': x_col},
            'yAxis': {'type': 'value', 'name': y_col},
            'series': [{'type': 'scatter', 'data': data, 'symbolSize': 8,
                        'itemStyle': {'color': '#5b8ff9', 'opacity': 0.7}}],
            'grid': {'bottom': 60, 'left': 70}
        }


# ==============================================================================
#  ExportManager — 模块 6: 导出
# ==============================================================================

class ExportManager:

    @staticmethod
    def to_csv_bytes(result):
        table = result.get('summary_table', [])
        columns = result.get('columns', [])
        if not table:
            return b''
        df = pd.DataFrame(table)
        if columns:
            df = df.reindex(columns=[c for c in columns if c in df.columns])
        buf = io.BytesIO()
        df.to_csv(buf, index=False, encoding='utf-8-sig')
        return buf.getvalue()

    @staticmethod
    def to_excel_bytes(result):
        table = result.get('summary_table', [])
        columns = result.get('columns', [])
        if not table:
            return b''
        df = pd.DataFrame(table)
        if columns:
            df = df.reindex(columns=[c for c in columns if c in df.columns])

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='统计结果')
            # 三线表样式
            ws = writer.sheets['统计结果']
            from openpyxl.styles import Font, Border, Side, Alignment
            thin = Side(style='thin')
            thick = Side(style='medium')
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, size=11)
                cell.border = Border(top=thick, bottom=thin)
                cell.alignment = Alignment(horizontal='center')
            last_row = len(df) + 1
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=last_row, column=col_idx)
                cell.border = Border(bottom=thick)
            # 脚注
            footnotes = result.get('footnotes', [])
            if footnotes:
                for i, fn in enumerate(footnotes):
                    ws.cell(row=last_row + 1 + i, column=1, value=fn).font = Font(size=9, italic=True)
        return buf.getvalue()

    @staticmethod
    def to_academic_html(result):
        """生成三线表 HTML"""
        table = result.get('summary_table', [])
        columns = result.get('columns', [])
        footnotes = result.get('footnotes', [])
        if not table or not columns:
            return '<p>无数据</p>'

        html = '<table class="academic-table">'
        html += '<thead><tr>'
        for c in columns:
            html += f'<th>{c}</th>'
        html += '</tr></thead><tbody>'
        for row in table:
            html += '<tr>'
            for c in columns:
                val = row.get(c, '')
                html += f'<td>{val}</td>'
            html += '</tr>'
        html += '</tbody></table>'
        if footnotes:
            html += '<div class="table-footnotes">'
            for fn in footnotes:
                html += f'<p>{fn}</p>'
            html += '</div>'
        return html


# ==============================================================================
#  工具函数
# ==============================================================================

def _flatten_dict(d, parent_key='', sep='.'):
    """递归展平嵌套字典，跳过 confidence 字段"""
    items = {}
    if not isinstance(d, dict):
        return items
    for k, v in d.items():
        if k in ('confidence',):
            continue
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(_flatten_dict(v, new_key, sep))
        elif isinstance(v, list):
            if v and isinstance(v[0], dict):
                for i, item in enumerate(v):
                    items.update(_flatten_dict(item, f"{new_key}[{i}]", sep))
            else:
                items[new_key] = str(v)
        else:
            items[new_key] = v
    return items
